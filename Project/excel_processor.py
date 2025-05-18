# import openpyxl
# import os
# import pandas as pd

# folder_path = os.path.join(os.path.dirname(__file__), "Excel inputs")

# all_months = [
#     "Apr-24", "May-24", "Jun-24", "Jul-24", "Aug-24", "Sep-24", "Oct-24", "Nov-24", "Dec-24",
#     "Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25",
#     "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26"
# ]

# def get_financial_year_range(month):
#     idx = all_months.index(month)
#     if "-24" in month or month == "Mar-25":
#         return all_months[:all_months.index("Mar-25") + 1]
#     return all_months[all_months.index("Apr-25"):all_months.index("Mar-26") + 1]

# def get_status_counts(ws):
#     completed = 0
#     pending = 0
#     for row in ws.iter_rows(min_row=2, max_col=1):
#         status = str(row[0].value).strip().lower() if row[0].value else ""
#         if status == "completed":
#             completed += 1
#         elif status == "pending":
#             pending += 1
#     return completed, pending

# def build_status_summary(input_month):
#     cumulative_completed = 0
#     cumulative_pending = 0
#     monthly_completed = 0
#     monthly_pending = 0
#     data_frames = []

#     fy_months = get_financial_year_range(input_month)

#     for filename in os.listdir(folder_path):
#         if filename.endswith(".xlsx"):
#             wb = openpyxl.load_workbook(os.path.join(folder_path, filename), data_only=True)
#             if input_month in wb.sheetnames:
#                 ws = wb[input_month]
#                 df = pd.DataFrame(ws.values)
#                 df.columns = df.iloc[0]
#                 df = df[1:]
#                 df["Source File"] = filename
#                 data_frames.append(df)

#                 comp, pend = get_status_counts(ws)
#                 cumulative_completed += comp
#                 cumulative_pending += pend
#                 monthly_completed += comp
#                 monthly_pending += pend

#     combined_df = pd.concat(data_frames, ignore_index=True) if data_frames else pd.DataFrame()
#     return combined_df, monthly_completed, monthly_pending, cumulative_completed, cumulative_pending

import openpyxl
import os

# Path to the folder containing Excel files
FOLDER_PATH = r'C:\Users\32002714\Desktop\Python\Preventive maintainance dashboard\Excel inputs'

# List of months in order
all_months = [
    "Apr-24", "May-24", "Jun-24", "Jul-24", "Aug-24", "Sep-24", "Oct-24", "Nov-24", "Dec-24",
    "Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25",
    "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26"
]

# Load all Excel workbooks in the folder
def load_workbooks():
    workbooks = {}
    for filename in os.listdir(FOLDER_PATH):
        if filename.endswith(".xlsx"):
            path = os.path.join(FOLDER_PATH, filename)
            wb = openpyxl.load_workbook(path, data_only=True)
            workbooks[filename] = wb
    return workbooks

# Extract completed and pending counts from a sheet
def get_status_counts(ws):
    completed = 0
    pending = 0
    for row in ws.iter_rows(min_row=2, max_col=1):  # only column A
        val = str(row[0].value).strip().lower() if row[0].value else ""
        if val == "completed":
            completed += 1
        elif val == "pending":
            pending += 1
    return completed, pending

# Determine the relevant financial year based on the month
def get_financial_year_range(month):
    idx = all_months.index(month)
    if "-24" in month or month == "Mar-25":
        return all_months[:all_months.index("Mar-25")+1]
    else:
        return all_months[all_months.index("Apr-25"):all_months.index("Mar-26")+1]

# Main function to get counts for dashboard
def build_status_summary(input_month):
    workbooks = load_workbooks()
    fy_months = get_financial_year_range(input_month)

    cumulative_completed = 0
    cumulative_pending = 0
    monthly_completed = 0
    monthly_pending = 0

    for month in fy_months:
        for wb in workbooks.values():
            if month in wb.sheetnames:
                ws = wb[month]
                comp, pend = get_status_counts(ws)
                cumulative_completed += comp
                cumulative_pending += pend
                if month == input_month:
                    monthly_completed = comp
                    monthly_pending = pend

    return {
        "monthly_completed": monthly_completed,
        "monthly_pending": monthly_pending,
        "cumulative_completed": cumulative_completed,
        "cumulative_pending": cumulative_pending,
    }

